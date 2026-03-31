"""
Excel export functionality for attendance reports.
Generates formatted Excel workbooks with attendance data.
"""

from io import BytesIO
from datetime import datetime
from typing import Optional, Dict, Any, List
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("⚠ openpyxl not installed - Excel export disabled")

logger = logging.getLogger(__name__)


class AttendanceExcelExporter:
    """
    Generates formatted Excel workbooks for attendance reports.
    Includes data validation, formatting, and charts.
    """

    def __init__(self):
        """Initialize Excel exporter."""
        self.available = OPENPYXL_AVAILABLE
        if OPENPYXL_AVAILABLE:
            logger.info("✓ Excel exporter initialized")

    def generate_attendance_excel(
        self,
        report: Dict[str, Any],
        class_doc: Optional[Dict[str, Any]] = None
    ) -> BytesIO:
        """
        Generate an Excel workbook from attendance report.

        Args:
            report: Attendance report dictionary
            class_doc: Class document for additional context

        Returns:
            BytesIO object containing Excel file

        Raises:
            RuntimeError: If openpyxl not available
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl library required for Excel export")

        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"

            # Set column widths
            ws.column_dimensions['A'].width = 20  # Student Name
            ws.column_dimensions['B'].width = 15  # Section
            ws.column_dimensions['C'].width = 18  # Engagement Time
            ws.column_dimensions['D'].width = 15  # Attendance %
            ws.column_dimensions['E'].width = 12  # Status

            # Define styles
            header_fill = PatternFill(
                start_color="366092",
                end_color="366092",
                fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            present_font = Font(color="006100", bold=True)

            absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            absent_font = Font(color="9C0006", bold=True)

            center_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Add title section
            ws['A1'] = "ATTENDANCE REPORT"
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
            ws.merge_cells('A1:E1')
            ws['A1'].alignment = center_alignment

            # Add class info
            row = 2
            if class_doc:
                ws[f'A{row}'] = f"Class: {class_doc.get('title', 'N/A')}"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

                ws[f'A{row}'] = f"Teacher: {class_doc.get('teacher_name', 'N/A')}"
                row += 1

            # Add report summary
            ws[f'A{row}'] = f"Date: {report.get('class_date', 'N/A')}"
            row += 1

            ws[f'A{row}'] = f"Total Students: {report.get('total_students', 0)}"
            row += 1

            ws[f'A{row}'] = f"Present: {report.get('present_count', 0)}"
            ws[f'A{row}'].font = Font(bold=True, color="006100")
            row += 1

            ws[f'A{row}'] = f"Absent: {report.get('absent_count', 0)}"
            ws[f'A{row}'].font = Font(bold=True, color="9C0006")
            row += 2

            # Add headers
            headers = ["Student Name", "Section", "Engagement Time", "Attendance %", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # Add data rows
            row += 1
            attendance_records = report.get('attendance_records', [])

            for record in attendance_records:
                # Student Name
                cell = ws.cell(row=row, column=1)
                cell.value = record.get('student_name', 'N/A')
                cell.border = border

                # Section
                cell = ws.cell(row=row, column=2)
                cell.value = record.get('section', 'N/A')
                cell.border = border
                cell.alignment = center_alignment

                # Engagement Time
                cell = ws.cell(row=row, column=3)
                cell.value = record.get('engagement_time_label', '0s')
                cell.border = border
                cell.alignment = center_alignment

                # Attendance %
                cell = ws.cell(row=row, column=4)
                engagement_pct = record.get('engagement_percentage', 0)
                cell.value = f"{engagement_pct:.1f}%"
                cell.border = border
                cell.alignment = center_alignment
                cell.number_format = '0.00"%"'

                # Status
                cell = ws.cell(row=row, column=5)
                status = record.get('attendance_status', 'absent').upper()
                cell.value = status
                cell.border = border
                cell.alignment = center_alignment

                # Apply status coloring
                if 'PRESENT' in str(status).upper():
                    cell.fill = present_fill
                    cell.font = present_font
                else:
                    cell.fill = absent_fill
                    cell.font = absent_font

                row += 1

            # Add summary statistics at bottom
            row += 1
            ws[f'A{row}'] = "Summary Statistics"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            row += 1

            # Attendance rate
            total_students = report.get('total_students', 0)
            present_count = report.get('present_count', 0)
            if total_students > 0:
                attendance_rate = (present_count / total_students) * 100
                ws[f'A{row}'] = f"Attendance Rate: {attendance_rate:.1f}%"
                row += 1

            # Class duration
            class_duration = report.get('class_duration_seconds', 0)
            if class_duration:
                minutes = class_duration // 60
                ws[f'A{row}'] = f"Class Duration: {minutes} minutes"
                row += 1

            # Started at
            started_at = report.get('started_at')
            if started_at:
                ws[f'A{row}'] = f"Started: {started_at}"
                row += 1

            # Ended at
            ended_at = report.get('ended_at')
            if ended_at:
                ws[f'A{row}'] = f"Ended: {ended_at}"
                row += 1

            # Add footer
            row += 2
            ws[f'A{row}'] = f"Report Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            ws[f'A{row}'].font = Font(italic=True, size=9)

            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            logger.info("✓ Excel attendance report generated successfully")
            return output

        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            raise

    def generate_bulk_attendance_excel(
        self,
        reports_list: List[Dict[str, Any]],
        class_doc: Optional[Dict[str, Any]] = None
    ) -> BytesIO:
        """
        Generate Excel workbook with multiple attendance reports (one sheet per report).

        Args:
            reports_list: List of attendance report dictionaries
            class_doc: Class document for context

        Returns:
            BytesIO object containing Excel file

        Raises:
            RuntimeError: If openpyxl not available
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl library required for Excel export")

        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Define styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            present_font = Font(color="006100", bold=True)

            absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            absent_font = Font(color="9C0006", bold=True)

            center_alignment = Alignment(horizontal="center", vertical="center")

            # Create a sheet for each report
            for idx, report in enumerate(reports_list, 1):
                session_id = report.get('session_id', f'Session_{idx}')
                # Limit sheet name to 31 chars (Excel limit)
                sheet_name = f"Class_{idx}" if len(session_id) > 20 else session_id[:20]

                ws = wb.create_sheet(title=sheet_name)

                # Set column widths
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 18
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 12

                # Title
                ws['A1'] = f"Class: {class_doc.get('title', 'N/A') if class_doc else 'N/A'}"
                ws['A1'].font = Font(bold=True, size=12)
                ws['A1'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                ws.merge_cells('A1:E1')

                # Date info
                row = 2
                ws[f'A{row}'] = f"Date: {report.get('class_date', 'N/A')}"
                row += 1

                ws[f'A{row}'] = f"Present: {report.get('present_count', 0)} | Absent: {report.get('absent_count', 0)}"
                row += 2

                # Headers
                headers = ["Student Name", "Section", "Engagement Time", "Attendance %", "Status"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border

                # Data rows
                row += 1
                for record in report.get('attendance_records', []):
                    ws.cell(row=row, column=1).value = record.get('student_name', 'N/A')
                    ws.cell(row=row, column=1).border = border

                    ws.cell(row=row, column=2).value = record.get('section', 'N/A')
                    ws.cell(row=row, column=2).border = border
                    ws.cell(row=row, column=2).alignment = center_alignment

                    ws.cell(row=row, column=3).value = record.get('engagement_time_label', '0s')
                    ws.cell(row=row, column=3).border = border
                    ws.cell(row=row, column=3).alignment = center_alignment

                    engagement_pct = record.get('engagement_percentage', 0)
                    ws.cell(row=row, column=4).value = f"{engagement_pct:.1f}%"
                    ws.cell(row=row, column=4).border = border
                    ws.cell(row=row, column=4).alignment = center_alignment

                    status = record.get('attendance_status', 'absent').upper()
                    cell = ws.cell(row=row, column=5)
                    cell.value = status
                    cell.border = border
                    cell.alignment = center_alignment

                    if 'PRESENT' in str(status).upper():
                        cell.fill = present_fill
                        cell.font = present_font
                    else:
                        cell.fill = absent_fill
                        cell.font = absent_font

                    row += 1

            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            logger.info(f"✓ Bulk Excel report generated with {len(reports_list)} sheets")
            return output

        except Exception as e:
            logger.error(f"Error generating bulk Excel report: {e}")
            raise


# Global exporter instance
attendance_excel_exporter = AttendanceExcelExporter()


def get_attendance_excel_exporter() -> AttendanceExcelExporter:
    """
    Get the global attendance Excel exporter instance.

    Returns:
        AttendanceExcelExporter instance
    """
    return attendance_excel_exporter
